#!/usr/bin/env python3
"""Build the combined Supplementary Data S1 workbook.

The program appends audited coffee and finger millet result tables to an existing
workbook. Inputs may be result directories or ZIP archives produced by the
analysis pipelines. Figure files are not read or embedded.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FILL = "0F5B5D"
SECTION_FILL = "DCEFEF"
SUBTLE_FILL = "F3F7F7"
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_GREY = Side(style="thin", color="C7D1D1")


@dataclass
class DataSource:
    path: Path

    def _zip_members(self) -> dict[str, str]:
        with zipfile.ZipFile(self.path) as archive:
            return {name.replace("\\", "/"): name for name in archive.namelist()}

    def find(self, suffixes: str | Sequence[str]) -> str:
        wanted = [suffixes] if isinstance(suffixes, str) else list(suffixes)
        wanted = [item.replace("\\", "/").lstrip("/") for item in wanted]
        if self.path.is_dir():
            candidates: list[tuple[str, Path]] = []
            for file_path in self.path.rglob("*"):
                if not file_path.is_file():
                    continue
                rel = file_path.relative_to(self.path).as_posix()
                for suffix in wanted:
                    if rel == suffix or rel.endswith("/" + suffix) or rel.endswith(suffix):
                        candidates.append((rel, file_path))
            if not candidates:
                raise FileNotFoundError(f"None of {wanted} was found under {self.path}")
            candidates.sort(key=lambda item: (len(item[0]), item[0]))
            return str(candidates[0][1])

        members = self._zip_members()
        candidates: list[tuple[str, str]] = []
        for normalized, original in members.items():
            for suffix in wanted:
                if normalized == suffix or normalized.endswith("/" + suffix) or normalized.endswith(suffix):
                    candidates.append((normalized, original))
        if not candidates:
            raise FileNotFoundError(f"None of {wanted} was found in {self.path}")
        candidates.sort(key=lambda item: (len(item[0]), item[0]))
        return candidates[0][1]

    def read_bytes(self, suffixes: str | Sequence[str]) -> bytes:
        selected = self.find(suffixes)
        if self.path.is_dir():
            return Path(selected).read_bytes()
        with zipfile.ZipFile(self.path) as archive:
            return archive.read(selected)

    def read_text(self, suffixes: str | Sequence[str]) -> str:
        return self.read_bytes(suffixes).decode("utf-8-sig", errors="replace")

    def table(self, suffixes: str | Sequence[str]) -> list[list[Any]]:
        selected = self.find(suffixes)
        suffix = Path(selected.replace("\\", "/")).suffix.lower()
        text = self.read_text(suffixes)
        delimiter = "\t" if suffix == ".tsv" else ","
        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
        return [[coerce_value(value) for value in row] for row in rows]

    def json(self, suffixes: str | Sequence[str]) -> Any:
        return json.loads(self.read_text(suffixes))


def coerce_value(value: str) -> Any:
    text = value.strip()
    if text == "":
        return None
    if text.lower() in {"true", "false"}:
        return text.lower() == "true"
    if re.fullmatch(r"[-+]?\d+", text):
        try:
            return int(text)
        except ValueError:
            return text
    if re.fullmatch(r"[-+]?(?:\d+\.\d*|\d*\.\d+)(?:[eE][-+]?\d+)?", text) or re.fullmatch(
        r"[-+]?\d+[eE][-+]?\d+", text
    ):
        try:
            return float(text)
        except ValueError:
            return text
    return sanitize_text(text)


def sanitize_text(value: str) -> str:
    text = value.replace("\\", "/")
    text = re.sub(r"(?i)^[A-Z]:/", "", text)
    text = re.sub(r"(?i)^.*/(?=(?:generated|results|analysis|reference)/)", "", text)
    return text


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", name)[:31]
    return cleaned or "Sheet"


def remove_sheet_if_present(workbook, name: str) -> None:
    if name in workbook.sheetnames:
        del workbook[name]


def write_table_sheet(workbook, name: str, rows: list[list[Any]], title: str, note: str = "") -> None:
    name = safe_sheet_name(name)
    remove_sheet_if_present(workbook, name)
    sheet = workbook.create_sheet(name)
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = title
    sheet["A1"].font = Font(size=14, bold=True, color="0F3C3D")
    if note:
        sheet["A2"] = note
        sheet["A2"].font = Font(italic=True, color="4A5A5A")
        sheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    start_row = 4 if note else 3
    if not rows:
        sheet.cell(start_row, 1, "No rows available")
        return

    max_cols = max(len(row) for row in rows)
    normalized = [list(row) + [None] * (max_cols - len(row)) for row in rows]
    for r_idx, row in enumerate(normalized, start=start_row):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(r_idx, c_idx, value)

    header = sheet[start_row]
    for cell in header:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GREY)
    sheet.freeze_panes = f"A{start_row + 1}"
    sheet.auto_filter.ref = f"A{start_row}:{get_column_letter(max_cols)}{start_row + len(normalized) - 1}"

    if (
        len(normalized) > 1
        and max_cols > 0
        and all(isinstance(value, str) and value.strip() for value in normalized[0])
    ):
        table_ref = f"A{start_row}:{get_column_letter(max_cols)}{start_row + len(normalized) - 1}"
        table_name = re.sub(r"[^A-Za-z0-9_]", "_", name) + "Table"
        table_name = ("T_" + table_name)[:240]
        try:
            table = Table(displayName=table_name, ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            sheet.add_table(table)
        except ValueError:
            pass

    for col_idx in range(1, max_cols + 1):
        values = [sheet.cell(row, col_idx).value for row in range(start_row, min(start_row + len(normalized), start_row + 250))]
        longest = max((len(str(value)) for value in values if value is not None), default=8)
        width = min(max(longest + 2, 10), 38)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    for row in sheet.iter_rows(min_row=start_row + 1, max_row=start_row + len(normalized) - 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=isinstance(cell.value, str) and len(cell.value) > 40)
            if isinstance(cell.value, float):
                cell.number_format = "0.000000"


def write_key_value_sheet(workbook, name: str, payload: Mapping[str, Any], title: str, note: str = "") -> None:
    rows: list[list[Any]] = [["Field", "Value"]]
    for key, value in flatten_mapping(payload):
        rows.append([humanize(key), value])
    write_table_sheet(workbook, name, rows, title, note)


def flatten_mapping(payload: Mapping[str, Any], prefix: str = "") -> Iterable[tuple[str, Any]]:
    for key, value in payload.items():
        full = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, Mapping):
            yield from flatten_mapping(value, full)
        elif isinstance(value, list):
            yield full, "; ".join(str(item) for item in value)
        else:
            yield full, value


def humanize(value: str) -> str:
    replacements = {
        "mixture_score_ot": "Topological Mixture Score",
        "negorc": "negative ORC incidence",
        "roc_auc": "ROC AUC",
        "average_precision": "average precision",
        "best_f1": "best F1",
        "knn": "kNN",
        "kmer": "k-mer",
        "r1": "R1",
        "r2": "R2",
    }
    text = value.replace("_", " ").replace(".", " — ")
    for source, target in replacements.items():
        text = re.sub(re.escape(source.replace("_", " ")), target, text, flags=re.IGNORECASE)
    return text[:1].upper() + text[1:]


def clean_existing_workbook(workbook) -> None:
    """Apply neutral labels to the distributed base workbook."""
    cell_updates = {
        ("Contents", "B2"): "Core coffee outputs and selected robustness tables organized in one workbook.",
        ("Contents", "B3"): "Local absolute paths are reduced to portable relative labels.",
        ("Contents", "B24"): "Idealized synthetic calibration parameters and performance summary.",
        ("Contents", "B32"): "Compact summary for kNN = 3–7 robustness.",
        ("Contents", "B40"): "Ablation analysis of Topological Mixture Score components.",
        ("KNN_Params", "A2"): "coffee_case_study_results",
    }
    for (sheet_name, address), value in cell_updates.items():
        if sheet_name in workbook.sheetnames:
            workbook[sheet_name][address] = value
    if "TMS_Ablation" in workbook.sheetnames:
        sheet = workbook["TMS_Ablation"]
        headers = [
            "Model / score definition", "N", "ROC AUC", "Average precision",
            "Best F1", "Best threshold", "Spearman rho with entropy",
            "Spearman P with entropy",
        ]
        labels = [
            "TMS (synthetic composite)",
            "Betweenness centrality only (Z-scored)",
            "Negative ORC incidence only (Z-scored)",
            "Composite: Z(betweenness) + Z(negative ORC)",
            "Composite (raw sum): betweenness + negative ORC",
        ]
        for column, value in enumerate(headers, start=1):
            sheet.cell(1, column, value)
        for row, value in enumerate(labels, start=2):
            sheet.cell(row, 1, value)


def add_cross_study_summary(workbook, metrics: list[list[Any]]) -> None:
    write_table_sheet(
        workbook,
        "CrossStudy_Summary",
        metrics,
        "Cross-study performance summary",
        "Primary and supporting endpoints from the coffee and finger millet benchmarks.",
    )


def rebuild_contents(workbook) -> None:
    name = "Contents"
    if name in workbook.sheetnames:
        del workbook[name]
    sheet = workbook.create_sheet(name, 0)
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "Supplementary Data S1"
    sheet["A1"].font = Font(size=16, bold=True, color="0F3C3D")
    sheet["A2"] = "Coffee case-study outputs, true read-level tests, and independent finger millet benchmarks."
    sheet["A2"].alignment = Alignment(wrap_text=True)
    rows = [["Worksheet", "Contents"]]
    descriptions = {
        "Analysis_Provenance": "Input archives, checksums, and build provenance.",
        "CrossStudy_Summary": "Primary performance metrics across studies and methods.",
    }
    for sheet_name in workbook.sheetnames:
        if sheet_name == "Contents":
            continue
        rows.append([sheet_name, descriptions.get(sheet_name, "Analysis output table.")])
    for r_idx, row in enumerate(rows, start=4):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(r_idx, c_idx, value)
    for cell in sheet[4]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = HEADER_FONT
    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 70
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:B{3 + len(rows)}"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--base_workbook", type=Path)
    parser.add_argument("--coffee_read_level", type=Path, required=True)
    parser.add_argument("--coffee_rare_event", type=Path, required=True)
    parser.add_argument("--coffee_reference", type=Path, required=True)
    parser.add_argument("--finger_design", type=Path, required=True)
    parser.add_argument("--finger_reference_free", type=Path, required=True)
    parser.add_argument("--finger_reference", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.base_workbook:
        workbook = load_workbook(args.base_workbook)
    else:
        workbook = Workbook()
        workbook.active.title = "Contents"
    clean_existing_workbook(workbook)

    c_read = DataSource(args.coffee_read_level)
    c_rare = DataSource(args.coffee_rare_event)
    c_ref = DataSource(args.coffee_reference)
    f_design = DataSource(args.finger_design)
    f_rf = DataSource(args.finger_reference_free)
    f_ref = DataSource(args.finger_reference)

    provenance = [["Input", "File", "SHA-256"]]
    for label, path in [
        ("Base workbook", args.base_workbook),
        ("Coffee read-level", args.coffee_read_level),
        ("Coffee rare-event", args.coffee_rare_event),
        ("Coffee reference", args.coffee_reference),
        ("Finger millet design", args.finger_design),
        ("Finger millet reference-free", args.finger_reference_free),
        ("Finger millet reference-based", args.finger_reference),
    ]:
        if path is None:
            continue
        provenance.append([label, path.name, sha256(path) if path.is_file() else "directory input"])
    write_table_sheet(workbook, "Analysis_Provenance", provenance, "Analysis provenance")

    # Coffee true read-level test
    write_key_value_sheet(workbook, "C_ReadLevel_Summary", c_read.json(["summary/read_level_master_metrics.json", "read_level_master_metrics.json"]), "Coffee true read-level mixture test")
    write_table_sheet(workbook, "C_ReadLevel_Runs", c_read.table("summary/run_summary.tsv"), "Coffee read-level run metrics")
    write_table_sheet(workbook, "C_ReadLevel_Comparators", c_read.table("summary/comparator_summary.tsv"), "Coffee read-level comparator metrics")
    write_table_sheet(workbook, "C_ReadLevel_Ratios", c_read.table("summary/ratio_summary.tsv"), "Coffee read-level mixture-ratio results")
    write_table_sheet(workbook, "C_ReadLevel_Truth", c_read.table("generated/manifests/truth_manifest.tsv"), "Coffee generated-library truth manifest")

    # Coffee rare-event test
    write_key_value_sheet(workbook, "C_RareEvent_Summary", c_rare.json(["summary/rare_event_master_metrics.json", "rare_event_master_metrics.json"]), "Coffee rare-event mixture-injection test")
    write_table_sheet(workbook, "C_RareEvent_Groups", c_rare.table(["summary/rare_event_group_summary.tsv", "rare_event_group_summary.tsv"]), "Coffee rare-event group summary")
    write_table_sheet(workbook, "C_RareEvent_Comparators", c_rare.table(["summary/rare_event_comparator_summary.tsv", "rare_event_comparator_summary.tsv"]), "Coffee rare-event comparator summary")
    write_table_sheet(workbook, "C_RareEvent_Patterns", c_rare.table(["summary/rare_event_pattern_summary.tsv", "rare_event_pattern_summary.tsv"]), "Coffee rare-event mixture-pattern summary")
    write_table_sheet(workbook, "C_RareEvent_ParentDist", c_rare.table(["summary/rare_event_parent_distance_strata.tsv", "rare_event_parent_distance_strata.tsv"]), "Coffee parent-distance strata")

    # Coffee leakage-controlled reference comparison
    write_key_value_sheet(workbook, "C_RefCrossfit_Summary", c_ref.json("summary/reference_qc_crossfit_master_metrics.json"), "Coffee leakage-controlled reference comparison")
    write_table_sheet(workbook, "C_RefCrossfit_Metrics", c_ref.table("comparison/crossfit_generated_consensus_metrics.tsv"), "Coffee cross-fitted generated-library metrics")
    write_table_sheet(workbook, "C_RefCrossfit_Folds", c_ref.table("scores/crossfit_generated_fold_provenance.tsv"), "Coffee cross-fitting fold provenance")
    write_table_sheet(workbook, "C_RefCrossfit_Real", c_ref.table("comparison/crossfit_real_consensus_and_tms.tsv"), "Coffee real-library score comparison")
    write_table_sheet(workbook, "C_Nominal_vs_Crossfit", c_ref.table("comparison/nominal_vs_crossfit_auc.tsv"), "Coffee nominal and cross-fitted AUC comparison")

    # Finger millet design and input QC
    write_table_sheet(workbook, "F_Panel83", f_design.table("source_selection/full_cohort_geometry_manifest_83.tsv"), "Finger millet 83-library panel")
    write_table_sheet(workbook, "F_QC83", f_design.table("qc/fastq_qc_per_sample.tsv"), "Finger millet paired-FASTQ quality summary")
    write_table_sheet(workbook, "F_Source28", f_design.table("source_selection/benchmark_source_panel_28.tsv"), "Finger millet 28-source benchmark panel")
    write_table_sheet(workbook, "F_ParentSets", f_design.table("design_lock/locked_parent_sets.tsv"), "Prespecified finger millet parent sets")
    write_table_sheet(workbook, "F_MixtureDesign84", f_design.table("design_lock/locked_mixture_definitions_84.tsv"), "Prespecified finger millet mixture definitions")
    write_table_sheet(workbook, "F_GeneratedDesign560", f_design.table("design_lock/locked_generated_library_design_560.tsv"), "Finger millet generated-library design")
    write_table_sheet(workbook, "F_Full83_JS", f_design.table("sketches/full83_js_distance.csv"), "Finger millet 83-library Jensen–Shannon distances")

    # Finger millet reference-free benchmark
    write_key_value_sheet(workbook, "F_RefFree_Summary", f_rf.json("summary/external_reference_free_master_metrics.json"), "Finger millet reference-free benchmark")
    write_table_sheet(workbook, "F_Full83_Nodes", f_rf.table("full_cohort/full83_node_scores.tsv"), "Finger millet full-cohort node scores")
    write_table_sheet(workbook, "F_Batch_Runs", f_rf.table("batch/batch_run_metrics.tsv"), "Finger millet batch benchmark metrics")
    write_table_sheet(workbook, "F_Batch_Categories", f_rf.table("summary/batch_category_summary.tsv"), "Finger millet batch category summary")
    write_table_sheet(workbook, "F_RareEvent_Summary", f_rf.table("summary/rare_event_group_summary.tsv"), "Finger millet reference-free rare-event summary")
    write_table_sheet(workbook, "F_RareEvent_Comparators", f_rf.table("summary/rare_event_comparator_summary.tsv"), "Finger millet reference-free comparator summary")

    # Finger millet reference-based benchmark
    write_key_value_sheet(workbook, "F_RefBased_Summary", f_ref.json("summary/external_reference_comparator_master_metrics.json"), "Finger millet reference-based benchmark")
    write_key_value_sheet(workbook, "F_Ref_MarkerSummary", f_ref.json("panel/marker_panel_summary.json"), "Finger millet independent marker-panel summary")
    write_table_sheet(workbook, "F_Ref_MarkerPanel", f_ref.table("panel/marker_panel.tsv"), "Finger millet independent marker panel")
    write_table_sheet(workbook, "F_Ref_CrossfitMetrics", f_ref.table("scores/generated_crossfit_metrics.tsv"), "Finger millet cross-fitted score metrics")
    write_table_sheet(workbook, "F_Ref_Folds", f_ref.table("scores/crossfit_fold_provenance.tsv"), "Finger millet cross-fitting fold provenance")
    write_table_sheet(workbook, "F_Ref_RareEvent", f_ref.table("summary/reference_comparator_rare_event_summary.tsv"), "Finger millet reference-based rare-event summary")

    add_cross_study_summary(
        workbook,
        [
            ["Dataset", "Evaluation", "Method", "ROC AUC", "Average precision", "Interpretation"],
            ["Coffee", "Idealized sketch-space calibration", "Reference-free TMS", 0.746154, 0.882480, "Controlled interpolation benchmark"],
            ["Coffee", "True read-level batch", "Reference-free TMS", 0.540705, 0.674352, "Near-chance ranking"],
            ["Coffee", "True read-level rare-event", "Reference-free TMS", 0.552885, None, "Not supported"],
            ["Coffee", "Leakage-controlled read-replicate holdout", "Reference-based QC", 0.527, None, "Not supported"],
            ["Finger millet", "True read-level batch", "Reference-free TMS", 0.493, None, "Near chance"],
            ["Finger millet", "One-mixture rare-event", "Reference-free TMS", 0.605, 0.159, "Weak-to-moderate enrichment"],
            ["Finger millet", "One-mixture rare-event", "Reference-based QC", 0.717517, 0.305, "Supported under disjoint-read holdout"],
        ],
    )

    rebuild_contents(workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"[DONE] {args.output}")


if __name__ == "__main__":
    main()
